# -*- coding: utf-8 -*-
"""mmconfig.py
"""
from openpyxl.styles import PatternFill, Font
from mmclass import USDMConfig, WordConfig, MdConfig

# USDM書式定義
REQCOLOR = PatternFill(patternType='solid', fgColor='e2efda') # 薄い緑
SUBCOLOR = PatternFill(patternType='solid', fgColor='ccccff') # 薄い青
REASONCOLOR = PatternFill(patternType='solid', fgColor='ffffcc') # 薄い黄
NONCOLOR = PatternFill(fill_type=None)
FBOLD = Font(bold=True)
NONBOLD = Font(bold=False)

# 各レベル項目定義
LV1 = 1
LV2 = 2
LV3 = 3
LV4 = 4
LVREASON = 5
LVREMARK = 6
LVSPEC = 7
LVNOTE = 8
LVNOTES = 9
LVNOTEE = 10

# レベル判別文字列
TXTREQ2 = '要求'
TXTREQ3 = 'サブ要求'
TXTREQ4 = '仕様グループ'
TXTSPEC = '仕様'
TXTREASON = '理由'
TXTREMARK = '備考'

usdm_cfg = USDMConfig(
    excel_file_ext='*.xlsx',
    usdm_file_pfx='_USDM.xlsx',
    usdm_sheet='USDM',
    tmpl_file='./templates/USDM_Template.xlsx',
    tmpl_sheet='Template',
    lv1_title_cell='B1',
    row_start=4,
    col_req=2,
    col_sub=3,
    col_grp=4,
    col_reason=5,
    col_remark=7,
    col_spec=5,
    col_module_start=9,
    freeze_panes='F4',
    prohibit_char=r"[ '*/:?[\]`’＊／：？［＼］￥]+",
    sub_txt='_',
    cell_style={
        TXTREQ2: [[NONCOLOR, NONBOLD], [REQCOLOR, FBOLD], [REQCOLOR, FBOLD], [REQCOLOR, FBOLD],
                  [REQCOLOR, FBOLD], [REQCOLOR, NONBOLD], [REQCOLOR, NONBOLD]],
        TXTREQ3: [[NONCOLOR, NONBOLD], [NONCOLOR, FBOLD], [SUBCOLOR, FBOLD], [SUBCOLOR, FBOLD],
                  [SUBCOLOR, FBOLD], [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD]],
        TXTREQ4: [[NONCOLOR, NONBOLD], [NONCOLOR, FBOLD], [NONCOLOR, FBOLD], [NONCOLOR, FBOLD],
                  [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD]],
        TXTSPEC: [[NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD],
                  [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD]],
        TXTREASON: [[NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD], [REASONCOLOR, NONBOLD],
                    [REASONCOLOR, NONBOLD], [NONCOLOR, NONBOLD], [NONCOLOR, NONBOLD]]
    },
    img_title_cell='B2',
    img_insert_cell='B4',
    img_title_font=FBOLD
)

word_cfg = WordConfig(
    tmpl_name='./templates/SpecTemplate.docx',
    file_name='仕様書_r0.docx',
    max_head_lv=4,
    head_lv_offset=1,
    txt_remark=TXTREMARK,
    txt_reason=TXTREASON
)

md_cfg = MdConfig(
    header='<link rel="stylesheet" href="style.css" />\n\n',
    section_remark='<section class="remark">\n',
    section_reason='<section class="reason">\n',
    sec_close='</section>\n\n',
    escape_strs=['\\', '*', '_', '#', '+', '-', '.', '!', '{', '}', '[', ']', '(', ')'],
    img_width='auto'
)

# Xmindファイル解析用
PTNLV = [
    [r'^(#|＃){1}\s+((\s|\S)+)', LV1],
    [r'^(#|＃){2}\s+((\s|\S)+)', LV2],
    [r'^(#|＃){3}\s+((\s|\S)+)', LV3],
    [r'^(#|＃){4}\s+((\s|\S)+)', LV4],
    [r'^(\?|？)\s+((\s|\S)+)', LVREASON],
    [r'^(\/\/)\s+((\s|\S)+)', LVREMARK],
    [r'^(!)\s*((\s|\S)*)', LVNOTE],
    [r'^(-->)\s*((\s|\S)*)', LVNOTEE],
    [r'^(<!--)\s*((\s|\S)*)', LVNOTES],
    [r'^((\s|\S)+)', LVSPEC],
]

# 画像保存設定
IMAGE_DIR = 'mm_images'
IMAGE_FILENAME_PROHIBIT = r'[\\/:*?"<>|\r\n\t]'
IMAGE_FILENAME_SUB = 'x'

# USDMフォーマットチェック
CHECKCELLS = ((3, 4, r'要求と仕様'), (3, 7, r'説明・備考'))
